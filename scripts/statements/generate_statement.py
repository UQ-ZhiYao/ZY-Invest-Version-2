#!/usr/bin/env python3
"""
Generate a ZY-Invest statement PDF from live Supabase data using the
finance-team's Excel templates, and file it in Supabase Storage + the
`statements` table.

Usage
-----
    python generate_statement.py subscription --tx-id <capital_injection.id>
    python generate_statement.py redemption   --tx-id <capital_injection.id>
    python generate_statement.py dividend     --investor-id <uuid> --fy-id <fy_settings.id>
    python generate_statement.py annual       --investor-id <uuid> --fy-id <fy_settings.id>

Requires SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY (env vars, or a .env file
next to this script — see .env.example). See README.md for the full setup,
what each subcommand does, and known data gaps.
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import sys
from pathlib import Path

SRC = Path(__file__).parent / "src"
sys.path.insert(0, str(SRC))

from compute import InvestorAddress, account_id, net_cost_asof, net_units_asof  # noqa: E402
from fill_annual import fill_annual_sheet  # noqa: E402
from fill_common import InvestorInfo, days_held_text  # noqa: E402
from fill_dividend import fill_dividend_sheet  # noqa: E402
from fill_subscription import fill_subscription_sheet  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from render_pdf import isolate_sheet_as_pdf  # noqa: E402
from storage import store_statement  # noqa: E402
from supa import Supabase  # noqa: E402

TEMPLATE_PATH = Path(__file__).parent / "templates" / "ZYInvest_Statement_Templates.xlsx"
OUT_DIR = Path(__file__).parent / "output"

DEFAULT_ACCOUNT_TYPE = "Direct Account"  # `profiles` has no account-type column yet — see README "Known gaps"
DEFAULT_SETTLEMENT_TYPE = "Banking"


def _load_dotenv() -> None:
    env_file = Path(__file__).parent / ".env"
    if not env_file.exists():
        return
    for line in env_file.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, val = line.partition("=")
        os.environ.setdefault(key.strip(), val.strip().strip('"').strip("'"))


def _investor_info(profile: dict, *, account_type_id: str | None = None,
                    issued_date: dt.date, asof: dt.date) -> InvestorInfo:
    addr = InvestorAddress.from_profile(profile)
    return InvestorInfo(
        account_type=DEFAULT_ACCOUNT_TYPE,
        account_id=account_type_id or account_id(DEFAULT_ACCOUNT_TYPE, issued_date),
        registered_name=profile.get("full_name") or "-",
        settlement_type=DEFAULT_SETTLEMENT_TYPE,
        phone=profile.get("phone") or "-",
        bank_name=profile.get("bank_name") or "-",
        email=profile.get("email") or "-",
        bank_account_no=profile.get("bank_account_no") or "-",
        nominee_or_joint_label="Nominee Name",  # `profiles` has no joint/nominee tracking yet
        nominee_or_joint_value="-",
        total_days_held_text=days_held_text(issued_date, asof),
        address_line1=addr.line1,
        address_line2=addr.line2,
        address_line3=addr.line3,
    )


def _profile_issued_date(profile: dict) -> dt.date:
    raw = profile.get("created_at") or dt.date.today().isoformat()
    return dt.datetime.fromisoformat(str(raw).replace("Z", "+00:00")).date()


def cmd_subscription_or_redemption(sb: Supabase, args) -> Path:
    tx = sb.select_one("capital_injection", {"id": f"eq.{args.tx_id}", "select": "*"})
    if not tx:
        raise SystemExit(f"No capital_injection row with id={args.tx_id}")
    if tx.get("status") != "Approved":
        print(f"warning: transaction status is '{tx.get('status')}', not 'Approved' — "
              "generating anyway, but consider approving it first.", file=sys.stderr)

    profile = sb.select_one("profiles", {"id": f"eq.{tx['uid']}", "select": "*"})
    if not profile:
        raise SystemExit(f"No profile found for uid={tx['uid']}")

    all_cis = sb.select("capital_injection", {"uid": f"eq.{tx['uid']}", "select": "*"})
    tx_date = dt.datetime.strptime(str(tx["date"])[:10], "%Y-%m-%d").date()
    prior = [r for r in all_cis if r["id"] != tx["id"]]
    opening_units = net_units_asof(prior, tx_date)
    opening_cost = net_cost_asof(prior, tx_date, tx["uid"])

    issued_date = _profile_issued_date(profile)
    investor = _investor_info(profile, issued_date=issued_date, asof=tx_date)

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Subscription"]
    fill_subscription_sheet(
        ws, tx=tx, investor=investor, opening_units=opening_units,
        opening_cost=opening_cost, issued_date=issued_date,
    )

    out_name = f"{tx['type']}_{tx.get('reference_id') or tx['id']}.pdf"
    out_path = OUT_DIR / out_name
    isolate_sheet_as_pdf(wb, "Subscription", out_path)

    if not args.no_upload:
        row = store_statement(
            sb, pdf_path=out_path, investor_id=tx["uid"], statement_type=tx["type"],
            period_label=tx_date.strftime("%d/%m/%Y"), transaction_id=tx["id"],
        )
        print(json.dumps(row, default=str, indent=2))
    return out_path


def cmd_dividend(sb: Supabase, args) -> Path:
    profile = sb.select_one("profiles", {"id": f"eq.{args.investor_id}", "select": "*"})
    if not profile:
        raise SystemExit(f"No profile found for id={args.investor_id}")
    fy = sb.select_one("fy_settings", {"id": f"eq.{args.fy_id}", "select": "*"})
    if not fy:
        raise SystemExit(f"No fy_settings row with id={args.fy_id}")

    dists = sb.select(
        "distributions",
        {"fy": f"eq.{fy['label']}", "select": "*", "order": "ex_date.asc"},
    )
    if not dists:
        raise SystemExit(f"No distributions found for FY '{fy['label']}'")

    all_cis = sb.select("capital_injection", {"uid": f"eq.{args.investor_id}", "select": "*"})
    fy_end = dt.datetime.strptime(str(fy["end_date"])[:10], "%Y-%m-%d").date()
    holding_units = net_units_asof(all_cis, fy_end, args.investor_id)

    issued_date = _profile_issued_date(profile)
    investor = _investor_info(profile, issued_date=issued_date, asof=fy_end)

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Dividend"]
    fill_dividend_sheet(
        ws, distributions=dists, investor=investor, holding_units=holding_units,
        period_text=fy["label"],
    )

    out_name = f"Dividend_{profile.get('full_name', 'investor').replace(' ', '_')}_{fy['label']}.pdf"
    out_path = OUT_DIR / out_name
    isolate_sheet_as_pdf(wb, "Dividend", out_path)

    if not args.no_upload:
        row = store_statement(
            sb, pdf_path=out_path, investor_id=args.investor_id, statement_type="Dividend",
            period_label=fy["label"], fy_id=fy["id"],
        )
        print(json.dumps(row, default=str, indent=2))
    return out_path


def cmd_annual(sb: Supabase, args) -> Path:
    profile = sb.select_one("profiles", {"id": f"eq.{args.investor_id}", "select": "*"})
    if not profile:
        raise SystemExit(f"No profile found for id={args.investor_id}")
    fy = sb.select_one("fy_settings", {"id": f"eq.{args.fy_id}", "select": "*"})
    if not fy:
        raise SystemExit(f"No fy_settings row with id={args.fy_id}")

    fy_start = dt.datetime.strptime(str(fy["start_date"])[:10], "%Y-%m-%d").date()
    fy_end = dt.datetime.strptime(str(fy["end_date"])[:10], "%Y-%m-%d").date()
    day_before_fy = fy_start - dt.timedelta(days=1)

    all_cis = sb.select("capital_injection", {"uid": f"eq.{args.investor_id}", "select": "*"})
    opening_units = net_units_asof(all_cis, day_before_fy, args.investor_id)
    opening_cost = net_cost_asof(all_cis, day_before_fy, args.investor_id)
    closing_units = net_units_asof(all_cis, fy_end, args.investor_id)
    closing_cost = net_cost_asof(all_cis, fy_end, args.investor_id)

    dists = sb.select("distributions", {"fy": f"eq.{fy['label']}", "select": "*", "order": "ex_date.asc"})

    nta_row = sb.select_one(
        "nta_daily", {"date": f"lte.{fy_end.isoformat()}", "select": "date,nta",
                      "order": "date.desc", "limit": "1"},
    )
    latest_nav = float(nta_row["nta"]) if nta_row else 1.0
    if not nta_row:
        print("warning: no nta_daily row on/before FY end — defaulting latest NAV to 1.0",
              file=sys.stderr)

    # Cashflow history for XIRR: subscriptions are outflows, redemptions are
    # inflows, dividends paid during the FY are inflows, and the closing
    # market value stands in for a final "sale" so the IRR reflects the
    # investor's current unrealized position too.
    cashflows: list[tuple[dt.date, float]] = []
    for r in all_cis:
        if r.get("status") != "Approved" or r.get("uid") != args.investor_id:
            continue
        d = dt.datetime.strptime(str(r["date"])[:10], "%Y-%m-%d").date()
        if d > fy_end:
            continue
        amt = float(r["amount"])
        cashflows.append((d, -amt if r["type"] == "Subscription" else amt))
    for d in dists:
        pay = d.get("pay_date") or d.get("ex_date")
        pay_date = dt.datetime.strptime(str(pay)[:10], "%Y-%m-%d").date()
        if fy_start <= pay_date <= fy_end:
            units_at_ex = net_units_asof(
                all_cis, dt.datetime.strptime(str(d["ex_date"])[:10], "%Y-%m-%d").date(),
                args.investor_id,
            )
            cashflows.append((pay_date, units_at_ex * float(d["dps"]) / 100.0))
    cashflows.append((fy_end, closing_units * latest_nav))

    issued_date = _profile_issued_date(profile)
    investor = _investor_info(profile, issued_date=issued_date, asof=fy_end)

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Annual"]
    fill_annual_sheet(
        ws, investor=investor, issued_date=issued_date, fy_start=fy_start, fy_end=fy_end,
        opening_units=opening_units, opening_cost=opening_cost,
        closing_units=closing_units, closing_cost=closing_cost,
        latest_nav_per_unit=latest_nav,
        distributions_in_fy=[d for d in dists if fy_start <= dt.datetime.strptime(
            str(d.get("pay_date") or d["ex_date"])[:10], "%Y-%m-%d").date() <= fy_end],
        cashflows_for_irr=cashflows,
        realized_pl=args.realized_pl, adjustment=args.adjustment,
    )

    out_name = f"Annual_{profile.get('full_name', 'investor').replace(' ', '_')}_{fy['label']}.pdf"
    out_path = OUT_DIR / out_name
    isolate_sheet_as_pdf(wb, "Annual", out_path)

    if not args.no_upload:
        row = store_statement(
            sb, pdf_path=out_path, investor_id=args.investor_id, statement_type="Annual",
            period_label=fy["label"], fy_id=fy["id"],
        )
        print(json.dumps(row, default=str, indent=2))
    return out_path


def main() -> None:
    _load_dotenv()
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = p.add_subparsers(dest="command", required=True)

    for name in ("subscription", "redemption"):
        sp = sub.add_parser(name, help=f"Generate a {name} statement for one capital_injection row")
        sp.add_argument("--tx-id", required=True)
        sp.add_argument("--no-upload", action="store_true", help="Write the PDF locally only, skip Storage/DB")
        sp.set_defaults(func=cmd_subscription_or_redemption)

    sp = sub.add_parser("dividend", help="Generate a dividend statement for one investor + FY")
    sp.add_argument("--investor-id", required=True)
    sp.add_argument("--fy-id", required=True)
    sp.add_argument("--no-upload", action="store_true")
    sp.set_defaults(func=cmd_dividend)

    sp = sub.add_parser("annual", help="Generate an annual investment account statement")
    sp.add_argument("--investor-id", required=True)
    sp.add_argument("--fy-id", required=True)
    sp.add_argument("--realized-pl", type=float, default=0.0,
                     help="Manual override — realized P&L isn't tracked in Supabase yet (see README)")
    sp.add_argument("--adjustment", type=float, default=0.0,
                     help="Manual one-off adjustment plug, same gap as the original template")
    sp.add_argument("--no-upload", action="store_true")
    sp.set_defaults(func=cmd_annual)

    args = p.parse_args()
    OUT_DIR.mkdir(exist_ok=True)
    sb = Supabase()
    out_path = args.func(sb, args)
    print(f"Wrote {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
