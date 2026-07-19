"""
Fills the "Annual" sheet (INVESTMENT ACCOUNT STATEMENT) for one investor over
one financial year.

Layout constraint: unlike Subscription/Dividend, this sheet is a fixed
2-page print layout — the Principal Transaction and Dividend Transaction
tables have a *fixed* number of rows (Opening/Closing only; 2 dividend
line-items) rather than one row per transaction, because inserting rows here
would shift the "Account Summary" block that shares the same row numbers in
a different column range and corrupt the merged-cell layout. So:

- Principal Transaction shows Opening (start of FY) and Closing (end of FY)
  balances, with the *net* cashflow/avg-cost/units-issued for the whole year
  summarised on the Closing row — not one row per subscription/redemption.
  If you need every transaction itemised, generate Subscription/Redemption
  statements for those individually (fill_subscription.py) alongside this one.
- Dividend Transaction supports up to 2 distributions in the FY (Interim +
  Final), same limit as fill_dividend.py.

Known gaps (not tracked in Supabase today, default to 0 / documented):
- Realized P&L: no ledger of realized gains on redemptions yet.
- "Adjustment": the template's own example uses this as a manual plug with
  no formula behind it — there's no source table for it, defaults to 0.
"""
from __future__ import annotations

import datetime as dt

from openpyxl.worksheet.worksheet import Worksheet

from compute import xirr
from fill_common import InvestorInfo, days_held_text, fill_investor_block_with_account_type, fill_page_header

FUND_EMAIL = "nzy.invest@gmail.com"
FUND_PHONE = "(+60)11 - 1121 8085"
MAX_DIVIDEND_ITEMS = 2


def fill_annual_sheet(
    ws: Worksheet,
    *,
    investor: InvestorInfo,
    issued_date: dt.date,
    fy_start: dt.date,
    fy_end: dt.date,
    opening_units: float,
    opening_cost: float,
    closing_units: float,
    closing_cost: float,
    latest_nav_per_unit: float,
    distributions_in_fy: list[dict],
    cashflows_for_irr: list[tuple[dt.date, float]],
    realized_pl: float = 0.0,
    adjustment: float = 0.0,
) -> None:
    if len(distributions_in_fy) > MAX_DIVIDEND_ITEMS:
        raise ValueError(
            f"Annual sheet only has {MAX_DIVIDEND_ITEMS} dividend line-item rows; "
            f"got {len(distributions_in_fy)} for this FY."
        )

    period_text = f"{fy_start.strftime('%d/%m/%Y')} - {fy_end.strftime('%d/%m/%Y')}"

    # --- header (page 1 + mirrored page-2 header) --------------------------
    ws["T1"] = None  # same stray '#VALUE!' artifact as A1, mirrored on the page-2 half
    fill_page_header(
        ws, statement_type="Annually", issued_date=dt.date.today(), period_text=period_text,
        fund_email=FUND_EMAIL, fund_phone=FUND_PHONE, investor=investor,
    )
    for col_src, col_dst in (("A8", "T8"), ("A9", "T9"), ("A10", "T10"), ("A11", "T11")):
        ws[col_dst] = ws[col_src].value
    for col_src, col_dst in (("O8", "AH8"), ("O9", "AH9"), ("O10", "AH10"), ("O11", "AH11"),
                              ("O12", "AH12")):
        ws[col_dst] = ws[col_src].value

    fill_investor_block_with_account_type(ws, investor)
    ws["N19"] = days_held_text(issued_date, fy_end)

    # --- Principal Transaction (T15:AI17) -----------------------------------
    net_cashflow = closing_cost - opening_cost
    net_units_issued = closing_units - opening_units
    avg_price = abs(net_cashflow / net_units_issued) if net_units_issued else None

    # The template left Z16/Z17/AC16/AC17 as 'General' format (it never shipped
    # example values there) — give them the same RM formatting used everywhere
    # else on the sheet instead of raw unformatted numbers.
    rm_2dp = '"RM"\\ #,##0.00'
    rm_4dp = '"RM"\\ 0.0000'

    ws["T16"] = fy_start
    ws["W16"] = "Opening"
    ws["Z16"] = "-"
    ws["AC16"] = round(opening_cost / opening_units, 4) if opening_units > 0 else "-"
    ws["AC16"].number_format = rm_4dp
    ws["AF16"] = "-"
    ws["AI16"] = round(opening_units, 4)

    ws["T17"] = fy_end
    ws["W17"] = "Closing"
    ws["Z17"] = round(net_cashflow, 2)
    ws["Z17"].number_format = rm_2dp
    ws["AC17"] = round(closing_cost / closing_units, 4) if closing_units > 0 else "-"
    ws["AC17"].number_format = rm_4dp
    ws["AF17"] = round(net_units_issued, 4)
    ws["AI17"] = round(closing_units, 4)

    # --- Dividend Transaction (T19:AI24) ------------------------------------
    ws["T21"] = fy_start
    ws["W21"] = "Opening"
    ws["AI21"] = 0

    running = 0.0
    for i in range(MAX_DIVIDEND_ITEMS):
        row = 22 + i
        if i < len(distributions_in_fy):
            d = distributions_in_fy[i]
            dps = float(d["dps"])
            amount = round(closing_units * dps / 100.0, 2)
            running += amount
            ws[f"T{row}"] = _parse_date(d["pay_date"] or d["ex_date"])
            ws[f"W{row}"] = f"{d.get('type') or ''} Dividend".strip()
            ws[f"AA{row}"] = round(dps, 4)
            ws[f"AC{row}"] = round(closing_units, 4)
            ws[f"AF{row}"] = amount
            ws[f"AI{row}"] = round(running, 2)
        else:
            for col in ("T", "W", "AA", "AC", "AF", "AI"):
                ws[f"{col}{row}"] = ""

    dividend_received = round(running, 2)
    ws["T24"] = fy_end
    ws["W24"] = "Closing"
    ws["AI24"] = dividend_received

    # --- Account Summary (A21:O31) -------------------------------------------
    market_value = round(closing_units * latest_nav_per_unit, 2)
    cost_basis = round(closing_cost, 2)
    unrealized_pl = round(market_value - cost_basis, 2)
    total_pl = round(unrealized_pl + realized_pl + dividend_received + adjustment, 2)
    total_perf_pct = (total_pl / cost_basis) if cost_basis else None

    ws["G23"] = round(closing_units, 4)
    ws["K23"] = round(latest_nav_per_unit, 6)
    ws["O23"] = market_value

    ws["G24"] = round(closing_units, 4)
    ws["K24"] = round(abs(cost_basis / closing_units), 6) if closing_units else "-"
    ws["O24"] = -cost_basis

    ws["O25"] = round(market_value - cost_basis, 2)  # (c) unrealized P&L
    ws["O26"] = round(realized_pl, 2)                 # (d) — see module docstring
    ws["O27"] = dividend_received                     # (e)
    ws["D28"] = "Adjustment"
    ws["O28"] = round(adjustment, 2)                  # (f) — see module docstring
    ws["O29"] = total_pl
    ws["O30"] = round(total_perf_pct, 4) if total_perf_pct is not None else "-"

    irr = xirr(cashflows_for_irr)
    ws["O31"] = round(irr, 4) if irr is not None else "-"


def _parse_date(v) -> dt.date:
    if isinstance(v, dt.date):
        return v
    return dt.datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
