"""
Fills the "Subscription" sheet for a single `capital_injection` row.

There is no separate Redemption sheet in the template, so a Redemption
transaction reuses this exact layout with the sign flipped (investment value
and units go down instead of up) and the title / column headers relabelled —
this matches how the admin console already treats Subscription/Redemption as
one table with a `type` column (assets/js/principal-admin.js).
"""
from __future__ import annotations

import datetime as dt

from openpyxl.worksheet.worksheet import Worksheet

from fill_common import (
    InvestorInfo,
    days_held_text,
    fill_investor_block_with_account_type,
    fill_page_header,
    force_single_page_landscape,
)

FUND_EMAIL = "nzy.invest@gmail.com"
FUND_PHONE = "(+60)11 - 1121 8085"


def fill_subscription_sheet(ws: Worksheet, *, tx: dict, investor: InvestorInfo,
                             opening_units: float, opening_cost: float,
                             issued_date: dt.date) -> None:
    tx_type = tx["type"]  # 'Subscription' or 'Redemption'
    is_redemption = tx_type == "Redemption"
    tx_date = _parse_date(tx["date"])
    amount = float(tx["amount"])
    price = float(tx["nta"])
    units = float(tx["units"])

    ws["K5"] = f"FUND  {tx_type.upper()}  STATEMENT"
    force_single_page_landscape(ws)

    fill_page_header(
        ws,
        statement_type=f"{tx_type} Statement",
        issued_date=dt.date.today(),
        period_text=tx_date.strftime("%d/%m/%Y"),
        fund_email=FUND_EMAIL,
        fund_phone=FUND_PHONE,
        investor=investor,
    )
    fill_investor_block_with_account_type(ws, investor)

    ws["G22"] = "Investment Value"
    ws["J22"] = f"{tx_type} Price"

    signed_amount = -amount if is_redemption else amount
    signed_units = -units if is_redemption else units
    closing_units = opening_units + signed_units
    closing_cost = opening_cost + signed_amount

    ws["A23"] = tx_date
    ws["D23"] = "Opening"
    ws["G23"] = round(opening_cost, 2)
    ws["J23"] = "-"
    ws["M23"] = round(opening_units, 4)
    ws["P23"] = round(opening_cost / opening_units, 4) if opening_units > 0 else "-"

    ws["A24"] = tx_date
    ws["D24"] = f"Fund {tx_type}"
    # G-column's number format hides negatives entirely ('"RM" #,##0.00;;"-"'
    # — an empty negative section), unlike the Units column which shows them
    # in red parens — so show the transaction's absolute amount here and let
    # the Description + signed Units/closing balance carry the direction.
    ws["G24"] = round(amount, 2)
    ws["J24"] = round(price, 6)
    ws["M24"] = round(signed_units, 4)
    ws["P24"] = "-"

    ws["A25"] = tx_date
    ws["D25"] = "Closing"
    ws["G25"] = round(closing_cost, 2)
    ws["J25"] = "-"
    ws["M25"] = round(closing_units, 4)
    ws["P25"] = round(closing_cost / closing_units, 4) if closing_units > 0 else "-"


def _parse_date(v) -> dt.date:
    if isinstance(v, dt.date):
        return v
    return dt.datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
