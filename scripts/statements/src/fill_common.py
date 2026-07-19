"""
Shared cell-writing helpers for the three statement sheets (Subscription,
Dividend, Annual) in templates/ZYInvest_Statement_Templates.xlsx.

Design note — values, not formulas
-----------------------------------
The template drives every field off one input cell (an investor's Registered
Name) via `XLOOKUP` against an "InvProfile" table on a hidden data sheet, so a
human can retype a name and watch the whole statement recompute. A generated
statement is different: it is a frozen, point-in-time record (like a bank
statement) that must NOT change if tomorrow's NAV or a later edit to the
investor's profile changes the lookup result. So instead of setting the
lookup-key cell and leaving formulas in place, every cell below is written as
a plain computed value, and — after filling — the sheet is lifted into a
sheet-only workbook with no other tabs, so nothing else in the file can still
reference live data.

Cell coordinates below were read directly from the shipped template and are
identical across the Subscription/Dividend/Annual sheets for the header
block, and identical between Subscription and Annual for the investor-info
block (Dividend's is one row shorter — it skips the Account Type/Account ID
row). See SCHEME_* below.
"""
from __future__ import annotations

import datetime as dt
from dataclasses import dataclass

from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class InvestorInfo:
    account_type: str
    account_id: str
    registered_name: str
    settlement_type: str
    phone: str
    bank_name: str
    email: str
    bank_account_no: str
    nominee_or_joint_label: str
    nominee_or_joint_value: str
    total_days_held_text: str
    address_line1: str
    address_line2: str
    address_line3: str


def clear_stray_error(ws: Worksheet) -> None:
    """A1 on every statement sheet holds a stray '#VALUE!' left over in the
    source template (a broken formula that was never cleaned up) — clear it."""
    ws["A1"] = None


def force_single_page_landscape(ws: Worksheet) -> None:
    """Subscription/Dividend are single logical pages, but the template's own
    page setup (portrait, no fit-to-width) splits their wide tables across
    two pages down the middle — fine for manual editing, not for a PDF
    that's meant to be handed to an investor. Force one landscape page.

    Not used for the Annual sheet: that one is a genuine 2-page statement
    with a manual column break already defined by the template's author
    (before column T) — forcing fit-to-1-page-wide there would cram both
    halves together instead of respecting that break.
    """
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def fill_page_header(ws: Worksheet, *, statement_type: str, issued_date: dt.date,
                      period_text: str, fund_email: str, fund_phone: str,
                      investor: InvestorInfo) -> None:
    """A8-A11 (address block) and L7:O12 (statement meta) — same coordinates
    on Subscription, Dividend and Annual."""
    clear_stray_error(ws)
    ws["A8"] = investor.registered_name.upper()
    ws["A9"] = investor.address_line1
    ws["A10"] = investor.address_line2
    ws["A11"] = investor.address_line3

    ws["O8"] = f":  {issued_date.strftime('%d-%m-%Y')}"
    ws["O9"] = f":  {statement_type}"
    ws["O10"] = f":  {period_text}"
    ws["O11"] = f":  {fund_email}"
    ws["O12"] = f":  {fund_phone}"


def fill_investor_block_with_account_type(ws: Worksheet, investor: InvestorInfo) -> None:
    """Subscription & Annual layout (A15:N19) — has the Account Type / Account ID row."""
    ws["E15"] = investor.account_type
    ws["N15"] = investor.account_id
    ws["E16"] = investor.registered_name
    ws["N16"] = investor.settlement_type
    ws["E17"] = investor.phone
    ws["N17"] = investor.bank_name
    ws["E18"] = investor.email
    ws["N18"] = investor.bank_account_no
    ws["A19"] = investor.nominee_or_joint_label
    ws["E19"] = investor.nominee_or_joint_value
    ws["N19"] = investor.total_days_held_text


def fill_investor_block_name_first(ws: Worksheet, investor: InvestorInfo) -> None:
    """Dividend layout (A15:N17) — starts at Registered Name, no Account Type/ID row."""
    ws["E15"] = investor.registered_name
    ws["N15"] = investor.settlement_type
    ws["E16"] = investor.phone
    ws["N16"] = investor.bank_name
    ws["E17"] = investor.email
    ws["N17"] = investor.bank_account_no


def days_held_text(issued: dt.date, asof: dt.date) -> str:
    return f"{(asof - issued).days:,}  days"
