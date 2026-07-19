"""
Fills the "Dividend" sheet for one investor over one or more `distributions`
rows (typically an Interim + Final pair within the same FY — that's the exact
shape the template ships with).

Known gap: EPS / DPR (Earnings Per Share / Dividend Payout Ratio) are columns
in the template but nothing in Supabase tracks per-instrument EPS today — they
render as "-" rather than a fabricated number. Add an `eps` column wherever
underlying instrument fundamentals get recorded if these need to be real.
"""
from __future__ import annotations

import datetime as dt

from openpyxl.worksheet.worksheet import Worksheet

from fill_common import (
    InvestorInfo,
    fill_investor_block_name_first,
    fill_page_header,
    force_single_page_landscape,
)

FUND_EMAIL = "nzy.invest@gmail.com"
FUND_PHONE = "(+60)11 - 1121 8085"

# The template ships with exactly 2 pre-formatted line-item rows (21, 22).
# More than that would need new merged-cell rows inserted into the sheet —
# out of scope here; generate one statement per FY (Interim+Final) instead.
MAX_LINE_ITEMS = 2
ROW_START = 21


def fill_dividend_sheet(ws: Worksheet, *, distributions: list[dict], investor: InvestorInfo,
                         holding_units: float, period_text: str) -> None:
    if not distributions:
        raise ValueError("at least one distribution row is required")
    if len(distributions) > MAX_LINE_ITEMS:
        raise ValueError(
            f"template only has {MAX_LINE_ITEMS} line-item rows; got {len(distributions)} "
            "distributions — generate separate statements or extend the sheet first."
        )

    latest_pay_date = max(_parse_date(d["pay_date"] or d["ex_date"]) for d in distributions)
    force_single_page_landscape(ws)

    fill_page_header(
        ws,
        statement_type="Dividend Statement",
        issued_date=dt.date.today(),
        period_text=period_text,
        fund_email=FUND_EMAIL,
        fund_phone=FUND_PHONE,
        investor=investor,
    )
    fill_investor_block_name_first(ws, investor)

    total_dps = 0.0
    total_amount = 0.0
    for i in range(MAX_LINE_ITEMS):
        row = ROW_START + i
        if i < len(distributions):
            d = distributions[i]
            dps = float(d["dps"])  # sen per unit
            amount = round(holding_units * dps / 100.0, 2)
            total_dps += dps
            total_amount += amount
            ws[f"A{row}"] = _parse_date(d["pay_date"] or d["ex_date"])
            # Match the template's own row style ("Interim Dividend") rather than
            # prefixing the FY — the FY is already shown as the Statement Period
            # above, and the longer text doesn't fit the column.
            ws[f"D{row}"] = f"{d.get('type') or ''} Dividend".strip()
            ws[f"G{row}"] = round(holding_units, 4)
            ws[f"J{row}"] = "-"  # EPS — not tracked yet, see module docstring
            ws[f"L{row}"] = "-"  # DPR — depends on EPS
            ws[f"N{row}"] = round(dps, 4)
            ws[f"P{row}"] = amount
        else:
            for col in ("A", "D", "G", "J", "L", "N", "P"):
                ws[f"{col}{row}"] = ""

    ws["N23"] = round(total_dps, 4)
    ws["P23"] = round(total_amount, 2)


def _parse_date(v) -> dt.date:
    if isinstance(v, dt.date):
        return v
    return dt.datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
