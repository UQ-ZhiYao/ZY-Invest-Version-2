#!/usr/bin/env python3
"""
Exercises the fill + PDF pipeline with synthetic data — no Supabase
connection needed. Run this after any change to the fill_*/render_pdf
modules to make sure the cell mapping still produces a valid PDF before
pointing the real CLI at production data.
"""
from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

SRC = Path(__file__).parent / "src"
sys.path.insert(0, str(SRC))

from compute import InvestorAddress, account_id, net_cost_asof, net_units_asof  # noqa: E402
from fill_annual import fill_annual_sheet  # noqa: E402
from fill_common import InvestorInfo, days_held_text  # noqa: E402
from fill_dividend import fill_dividend_sheet  # noqa: E402
from fill_subscription import fill_subscription_sheet  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from render_pdf import isolate_sheet_as_pdf  # noqa: E402

TEMPLATE_PATH = Path(__file__).parent / "templates" / "ZYInvest_Statement_Templates.xlsx"
OUT_DIR = Path(__file__).parent / "output" / "offline_test"

PROFILE = {
    "id": "11111111-1111-1111-1111-111111111111",
    "full_name": "Ng Mon Teng",
    "phone": "(+60)12 - 7188 288",
    "email": "osaka88@gmail.com",
    "address": "No.19 Jalan Rusa, Taman Mohammad Yassin, 86200 Simpang Renggam",
    "bank_name": "RHB Bank",
    "bank_account_no": "2013 1630 0173 90",
    "created_at": "2023-08-17T00:00:00+00:00",
}

CIS = [
    {"id": "a1", "uid": PROFILE["id"], "date": "2023-08-17", "type": "Subscription",
     "amount": 30000.0, "nta": 1.0, "units": 30000.0, "status": "Approved"},
    {"id": "a2", "uid": PROFILE["id"], "date": "2025-06-01", "type": "Subscription",
     "amount": 8000.0, "nta": 0.994033683664741, "units": 8048.06, "status": "Approved"},
    {"id": "a3", "uid": PROFILE["id"], "date": "2025-07-10", "type": "Redemption",
     "amount": 5000.0, "nta": 1.05, "units": 4761.9, "status": "Approved"},
]

DISTS = [
    {"id": "d1", "fy": "2025", "type": "Interim", "ex_date": "2025-06-12", "pay_date": "2025-06-20",
     "dps": 2.1, "units": 200000, "status": "Paid"},
    {"id": "d2", "fy": "2025", "type": "Final", "ex_date": "2025-10-05", "pay_date": "2025-10-15",
     "dps": 1.0, "units": 200000, "status": "Paid"},
]

FY = {"id": "fy2025", "label": "2025", "start_date": "2024-12-01", "end_date": "2025-11-30"}


def _investor_info(asof: dt.date) -> InvestorInfo:
    addr = InvestorAddress.from_profile(PROFILE)
    issued = dt.date(2023, 8, 17)
    return InvestorInfo(
        account_type="Direct Account",
        account_id=account_id("Direct Account", issued),
        registered_name=PROFILE["full_name"],
        settlement_type="Banking",
        phone=PROFILE["phone"],
        bank_name=PROFILE["bank_name"],
        email=PROFILE["email"],
        bank_account_no=PROFILE["bank_account_no"],
        nominee_or_joint_label="Nominee Name",
        nominee_or_joint_value="-",
        total_days_held_text=days_held_text(issued, asof),
        address_line1=addr.line1, address_line2=addr.line2, address_line3=addr.line3,
    )


def test_subscription():
    tx = CIS[1]  # the 2025-06-01 subscription
    tx_date = dt.date(2025, 6, 1)
    prior = [r for r in CIS if r["id"] != tx["id"]]
    opening_units = net_units_asof(prior, tx_date)
    opening_cost = net_cost_asof(prior, tx_date, tx["uid"])

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Subscription"]
    fill_subscription_sheet(
        ws, tx=tx, investor=_investor_info(tx_date), opening_units=opening_units,
        opening_cost=opening_cost, issued_date=dt.date(2023, 8, 17),
    )
    out = isolate_sheet_as_pdf(wb, "Subscription", OUT_DIR / "subscription.pdf")
    print("wrote", out, out.stat().st_size, "bytes")


def test_redemption():
    tx = CIS[2]
    tx_date = dt.date(2025, 7, 10)
    prior = [r for r in CIS if r["id"] != tx["id"]]
    opening_units = net_units_asof(prior, tx_date)
    opening_cost = net_cost_asof(prior, tx_date, tx["uid"])

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Subscription"]
    fill_subscription_sheet(
        ws, tx=tx, investor=_investor_info(tx_date), opening_units=opening_units,
        opening_cost=opening_cost, issued_date=dt.date(2023, 8, 17),
    )
    out = isolate_sheet_as_pdf(wb, "Subscription", OUT_DIR / "redemption.pdf")
    print("wrote", out, out.stat().st_size, "bytes")


def test_dividend():
    fy_end = dt.date(2025, 11, 30)
    holding_units = net_units_asof(CIS, fy_end, PROFILE["id"])
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Dividend"]
    fill_dividend_sheet(
        ws, distributions=DISTS, investor=_investor_info(fy_end),
        holding_units=holding_units, period_text=FY["label"],
    )
    out = isolate_sheet_as_pdf(wb, "Dividend", OUT_DIR / "dividend.pdf")
    print("wrote", out, out.stat().st_size, "bytes")


def test_annual():
    fy_start, fy_end = dt.date(2024, 12, 1), dt.date(2025, 11, 30)
    day_before = fy_start - dt.timedelta(days=1)
    opening_units = net_units_asof(CIS, day_before, PROFILE["id"])
    opening_cost = net_cost_asof(CIS, day_before, PROFILE["id"])
    closing_units = net_units_asof(CIS, fy_end, PROFILE["id"])
    closing_cost = net_cost_asof(CIS, fy_end, PROFILE["id"])

    cashflows = []
    for r in CIS:
        d = dt.datetime.strptime(r["date"], "%Y-%m-%d").date()
        if d > fy_end:
            continue
        amt = r["amount"]
        cashflows.append((d, -amt if r["type"] == "Subscription" else amt))
    dists_in_fy = [d for d in DISTS
                   if fy_start <= dt.datetime.strptime(d["pay_date"], "%Y-%m-%d").date() <= fy_end]
    for d in dists_in_fy:
        pay_date = dt.datetime.strptime(d["pay_date"], "%Y-%m-%d").date()
        ex_date = dt.datetime.strptime(d["ex_date"], "%Y-%m-%d").date()
        units_at_ex = net_units_asof(CIS, ex_date, PROFILE["id"])
        cashflows.append((pay_date, units_at_ex * d["dps"] / 100.0))
    latest_nav = 1.11254828669048
    cashflows.append((fy_end, closing_units * latest_nav))

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Annual"]
    fill_annual_sheet(
        ws, investor=_investor_info(fy_end), issued_date=dt.date(2023, 8, 17),
        fy_start=fy_start, fy_end=fy_end,
        opening_units=opening_units, opening_cost=opening_cost,
        closing_units=closing_units, closing_cost=closing_cost,
        latest_nav_per_unit=latest_nav, distributions_in_fy=dists_in_fy,
        cashflows_for_irr=cashflows,
    )
    out = isolate_sheet_as_pdf(wb, "Annual", OUT_DIR / "annual.pdf")
    print("wrote", out, out.stat().st_size, "bytes")


if __name__ == "__main__":
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    test_subscription()
    test_redemption()
    test_dividend()
    test_annual()
